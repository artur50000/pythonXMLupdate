import os
import shutil
from openpyxl import Workbook
import re

directory = path_to_directory
for filename in os.listdir(directory):
    if filename.endswith(".xml"):
        newpath = shutil.copy(
            directory + "/" + filename,
            path_to_new_directory)
        os.remove(directory + "/" + filename)


wb = Workbook()
ws = wb.active
i = 2

p = re.compile('>(.*)<')

# parsing of xml file

directory1 = path_to_new_directory
for filename in os.listdir(directory1):
    f = open(directory1 + "/" + filename)
    lst = f.readlines()
    lst2 = []
    update = filename.replace(".xml", "")
    date1 = ''

    for line in lst:
        if 'CreateDate=' in line:

            date1 = line[line.index('CreateDate=') +
                         12:line.index('CreateDate=') + 23]

    for k in range(len(lst)):
        if update in lst[k]:
            lst2.append(lst[k:k + 8])

    for item in lst2:
        ws["B" + str(i)] = update

        for line in item:

            ws["A" + str(i)] = date1

            if '<Reference>' in line:

                result = p.search(line)

                ws["C" + str(i)] = result.group(1)

            if '<symptom>' in line:

                result = p.search(line)

                ws["D" + str(i)] = result.group(1)

            if '<Problem>' in line:

                result = p.search(line)

                ws["E" + str(i)] = result.group(1)

            if '<NatureOfFix>' in line:

                result = p.search(line)

                ws["F" + str(i)] = result.group(1)
        i += 1

    f.close()

print(str(i) + " lines recorded")
wb.save(path_to_excel_file + "testvictorypython.xls")
